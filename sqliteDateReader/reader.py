from sqlite3 import connect
from os.path import exists
from openpyxl import Workbook
from tqdm import tqdm

def toExcel(colname, data):
    book = Workbook()
    sheet = book.create_sheet(index=0)


    for index, col in enumerate(colname):
        sheet.cell(1, index+1, col)
    for index, row in tqdm(enumerate(data)):
        for index2, column in enumerate(row):
            sheet.cell(index+2, index2+1, column)
    book.save(filename+".xlsx")

def readDatabase():
    global filename
    conn = connect(filename)
    c = conn.cursor()
    colname = [i[0] for i in c.execute("select * from hisdata").description]
    data = c.execute("select * from hisdata").fetchall()
    toExcel(colname, data)

def getFilename():
    global filename
    while True:
        filename = input("请输入数据库文件名称：")
        if exists(filename): break
        else: print("文件不存在，请重新输入。")

def main():
    getFilename()
    readDatabase()

if __name__ == '__main__':
    filename = ''
    main()